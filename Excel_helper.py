import openpyxl
from openpyxl import load_workbook
import os
import subprocess
import platform
import Global_variables

def populate_invoice_numbers(
        file_path: str,
        sheet_name: str,
        column_name: str,
        invoice_numbers: list
):
    """
    Populates the 'Invoice Number' column in an Excel sheet with values from 'invoice_numbers'.
    If a row already has a value in that column, it is skipped.
    :param file_path: Full path to the Excel file (e.g., r"C:\...\Failed Invoices\invoices.xlsx").
    :param sheet_name: The sheet name in the workbook where the data is located.
    :param column_name: The header name for the 'Invoice Number' column.
    :param invoice_numbers: A list of integers or strings representing invoice numbers.
    """

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Find which column index corresponds to 'column_name' in the first row
    col_index = None
    for cell in ws[1]:  # Assuming the first row has the headers
        if cell.value == column_name:
            col_index = cell.column
            break

    if col_index is None:
        raise ValueError(f"Column '{column_name}' not found in the header row.")

    # Iterate over rows starting from row 2 (to skip the header)
    # and fill in the invoice numbers in any empty cells under the specified column
    num_index = 0
    for row in range(2, 1000):
        # Stop if we've assigned all invoice numbers
        if num_index >= len(invoice_numbers):
            break
        current_cell = ws.cell(row=row, column=col_index)

        # If the cell is empty, assign the next invoice number
        if current_cell.value is None or current_cell.value == "":

            current_cell.value = invoice_numbers[num_index]
            num_index += 1

    # Save changes
    wb.save(file_path)

def read_column_values(file_path: str, sheet_name: str, column_name: str) -> list:
    """
    Reads and returns all values from the specified column in an Excel sheet.

    :param file_path: Full path to the Excel file.
    :param sheet_name: The name of the worksheet to read from.
    :param column_name: The exact header of the target column in the first row.
    :return: A list of values from the specified column (excluding the header).
    """
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Select the worksheet
    ws = wb[sheet_name]

    # Identify which column index corresponds to the given column_name
    # Assuming the first row is the header row
    column_index = None
    for cell in ws[1]:
        if cell.value == column_name:
            column_index = cell.column
            break

    if column_index is None:
        raise ValueError(f"Column '{column_name}' not found in the first row headers.")

    # Read all values from the target column, excluding the header (first row)
    values = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_index).value
        if cell_value is None:
            continue
        values.append(cell_value)

    return values

def insert_tuples_in_excel(file_path: str, sheet_name: str, data: list[tuple]) -> None:
    """
    Inserts each tuple in 'data' into the first empty rows in the given Excel sheet.
    If a row already has any content (in any column), that row is skipped.

    :param file_path: Full path to the Excel file.
    :param sheet_name: The worksheet name to write to.
    :param data: A list of tuples. Each tuple's items are placed into one row, in consecutive columns.
    """
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Keep track of which tuple we are on
    data_index = 0
    total_tuples = len(data)

    # We will check rows starting from 2 (to skip header row),
    # and go up to a "safe" upper limit (current max_row + number of tuples).
    max_possible_row = ws.max_row + total_tuples + 10  # a buffer of 10 rows, adjust if needed

    # Iterate over each row until we place all tuples
    for row in range(2, max_possible_row + 1):
        if data_index >= total_tuples:
            break  # we've placed all the tuples

        # Check if this row is empty in *any* column up to ws.max_column
        # You could expand this range if you want to be sure about columns beyond max_column.
        row_has_content = any(ws.cell(row=row, column=col).value
                              for col in range(1, ws.max_column + 1))

        if not row_has_content:
            # Row is empty, so place the tuple here
            current_tuple = data[data_index]
            for col_index, value in enumerate(current_tuple, start=1):
                ws.cell(row=row, column=col_index, value=value)
            data_index += 1
        else:
            # This row already has content, skip it
            continue

    # Save the workbook
    wb.save(file_path)


def delete_cell_content_if_matches(
        file_path: str,
        sheet_name: str,
        column_name: str,
        match_value
) -> None:
    """
    Deletes cell content in the specified column if it matches 'match_value'.

    :param file_path: Full path to the Excel file.
    :param sheet_name: Name of the worksheet in which the column resides.
    :param column_name: The exact header of the column (as it appears in the first row).
    :param match_value: The value to match. If the cell's value equals this, it will be cleared.
    """

    # Load the workbook and select the specified sheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Identify which column index corresponds to 'column_name' in the header row
    col_index = None
    for cell in ws[1]:  # Assuming the first row has column headers
        if cell.value == column_name:
            col_index = cell.column
            break

    if col_index is None:
        raise ValueError(f"Column '{column_name}' not found in the header row.")

    # Iterate over all rows after the header to find matching values
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_index)
        if cell.value == match_value:
            # Clear the cell content
            cell.value = None
            # Clear the cell with col index+1
            cell = ws.cell(row=row, column=col_index+1)
            cell.value = None

    # Save changes
    wb.save(file_path)


def clear_all(
        file_path: str,
        sheet_name: str
) -> None:
    """
    Removes all content from the specified sheet in an Excel file,
    keeping only the first row (header). Overwrites the original file.

    :param file_path: Path to the Excel file (e.g., 'data.xlsx').
    :param sheet_name: Name of the sheet to clear (e.g., 'Sheet1').
    """
    # Load the workbook and select the given sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # If there are more than 1 row, remove everything from the second row onward
    if sheet.max_row > 1:
        # delete_rows(start_row, number_of_rows)
        # Here we start at row 2, and delete (max_row - 1) rows
        sheet.delete_rows(2, sheet.max_row - 1)

    # Save changes back to the same file
    workbook.save(file_path)


def read_cell_content_from_first_two_col(
        file_path: str,
        sheet_name: str
) -> list[tuple]:
    """
    Reads the values from the first two columns of the given Excel sheet.
    Skips the first row (treated as header).

    Returns:
        A list of tuples, each containing the values (col1, col2) for one row.
    """
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    data = []
    # Start from row=2 to skip the header (row=1).
    # Read only columns 1 and 2.
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        # row is a tuple like (value_in_col1, value_in_col2)
        if row[0] is None:
            continue
        data.append((row[0], row[1]))

    return data


def open_excel_app(file_path: str):
    """
    Opens an Excel file in the native Microsoft Excel application

    Parameters:
    file_path (str): Path to the Excel file (.xlsx, .xls)

    Returns:
    bool: True if successful, False otherwise
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        system = platform.system()

        if system == "Windows":
            # For Windows
            os.startfile(file_path)
            return True

        elif system == "Darwin":
            # For macOS
            subprocess.run(["open", file_path], check=True)
            return True

        elif system == "Linux":
            # For Linux
            subprocess.run(["xdg-open", file_path], check=True)
            return True

        else:
            raise OSError("Unsupported operating system")

    except Exception as e:
        print(f"Error opening Excel: {e}")
        return False

def open_succeed_invoices():
    open_excel_app(Global_variables.secceed_invoices_excel_path)

def open_funding_requested_invoices():
    open_excel_app(Global_variables.funding_requested_excel_path)

def open_failed_invoices():
    open_excel_app(Global_variables.failed_invoices_excel_path)